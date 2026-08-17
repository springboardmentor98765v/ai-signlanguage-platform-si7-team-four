import csv
import os
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from openpyxl import Workbook

from typing import Any, Dict, List
from sqlalchemy.orm import Session

from app.models.models import (
    User,
    AnalyticsSummary,
    Certificate,
    Assessment,
    PracticeSession,
)

from app.services import streak_service
from app.services.analytics_service import get_learner_analytics


def _get_analytics_summary(
    db: Session,
    learner_id: str,
):
    return (
        db.query(AnalyticsSummary)
        .filter(
            AnalyticsSummary.user_id == learner_id
        )
        .first()
    )


def _count_lessons_completed(raw_data: Dict[str, Any]) -> int:
    """
    Count the number of lessons a learner has completed.

    Args:
        raw_data: Raw learner data as returned by _fetch_learner_raw_data.

    Returns:
        Number of completed lessons.
    """
    return len(raw_data.get("completed_lessons", []))


def _calculate_average_score(raw_data: Dict[str, Any]) -> float:
    """
    Calculate the learner's average score across all completed lessons.

    Args:
        raw_data: Raw learner data as returned by _fetch_learner_raw_data.

    Returns:
        Average score rounded to 2 decimal places. Returns 0.0 if the
        learner has no completed lessons.
    """
    lessons = raw_data.get("completed_lessons", [])

    if not lessons:
        return 0.0

    total_score = sum(
        lesson["score"]
        for lesson in lessons
    )

    return round(
        total_score / len(lessons),
        2
    )


def _identify_weak_letters(
    raw_data: Dict[str, Any],
    accuracy_threshold: float = 0.7,
) -> List[str]:
    """
    Identify letters the learner struggles with, based on accuracy.

    Args:
        raw_data: Raw learner data as returned by _fetch_learner_raw_data.
        accuracy_threshold: Minimum accuracy (0.0 - 1.0) required for a
            letter to NOT be considered weak. Defaults to 0.7 (70%).

    Returns:
        A list of letters (strings) identified as weak, sorted
        alphabetically.
    """
    weak_letters: List[str] = []

    letter_attempts = raw_data.get(
        "letter_attempts",
        {}
    )

    for letter, stats in letter_attempts.items():
        attempts = stats.get(
            "attempts",
            0
        )

        correct = stats.get(
            "correct",
            0
        )

        if attempts == 0:
            continue

        accuracy = correct / attempts

        if accuracy < accuracy_threshold:
            weak_letters.append(letter)

    return sorted(weak_letters)


def _list_certificates_earned(
    raw_data: Dict[str, Any],
) -> List[str]:
    """
    Extract the titles of certificates earned by the learner.

    Args:
        raw_data: Raw learner data as returned by _fetch_learner_raw_data.

    Returns:
        A list of certificate titles earned by the learner.
    """
    certificates = raw_data.get(
        "certificates",
        []
    )

    return [
        cert["title"]
        for cert in certificates
    ]


def generate_progress_report(
    db: Session,
    learner_id: str,
) -> Dict:

    analytics = _get_analytics_summary(
        db,
        learner_id,
    )

    if analytics is None:
        raise ValueError("Learner not found.")

    certificates = (
        db.query(Certificate)
        .filter(
            Certificate.user_id == learner_id
        )
        .count()
    )

    streak = streak_service.get_user_streak(
        db,
        learner_id,
    )

    return {
        "learner_id": learner_id,
        "lessons_completed": analytics.lessons_completed,
        "average_accuracy": analytics.overall_accuracy_percentage,
        "practice_hours": analytics.practice_hours,
        "improvement_rate": analytics.improvement_rate_percentage,
        "current_streak": streak["current_streak"],
        "certificates_earned": certificates,
    }


def export_progress_report_csv(
    db: Session,
    learner_id: str,
) -> str:
    """
    Exports the learner's progress report as a CSV file.

    Args:
        learner_id: Unique identifier of the learner.

    Returns:
        Path to the generated CSV file.
    """

    report = generate_progress_report(
        db,
        learner_id,
    )

    os.makedirs(
        "reports",
        exist_ok=True
    )

    file_path = os.path.join(
        "reports",
        f"progress_report_{learner_id}.csv"
    )

    with open(
        file_path,
        mode="w",
        newline="",
        encoding="utf-8",
    ) as csv_file:

        writer = csv.writer(csv_file)

        writer.writerow(
            ["Field", "Value"]
        )

        writer.writerow(
            ["Learner ID", report["learner_id"]]
        )

        writer.writerow(
            ["Lessons Completed", report["lessons_completed"]]
        )

        writer.writerow(
            ["Average Accuracy", report["average_accuracy"]]
        )

        writer.writerow(
            ["Practice Hours", report["practice_hours"]]
        )

        writer.writerow(
            ["Improvement Rate", report["improvement_rate"]]
        )

        writer.writerow(
            ["Current Streak", report["current_streak"]]
        )

        writer.writerow(
            [
                "Certificates Earned",
                report["certificates_earned"],
            ]
        )

    return file_path


def export_progress_report_pdf(
    db: Session,
    learner_id: str,
) -> str:
    report = generate_progress_report(
        db,
        learner_id,
    )

    os.makedirs(
        "reports",
        exist_ok=True
    )

    file_path = os.path.join(
        "reports",
        f"progress_report_{learner_id}.pdf"
    )

    document = SimpleDocTemplate(file_path)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Progress Report", styles["Title"]))
    story.append(Spacer(1, 20))

    rows = [
        ["Field", "Value"],
        ["Learner ID", report["learner_id"]],
        ["Lessons Completed", report["lessons_completed"]],
        ["Average Accuracy", report["average_accuracy"]],
        ["Practice Hours", report["practice_hours"]],
        ["Improvement Rate", report["improvement_rate"]],
        ["Current Streak", report["current_streak"]],
        ["Certificates Earned", report["certificates_earned"]],
    ]

    table = Table(rows)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    story.append(table)
    document.build(story)

    return file_path


def export_progress_report_excel(
    db: Session,
    learner_id: str,
) -> str:
    report = generate_progress_report(
        db,
        learner_id,
    )

    os.makedirs(
        "reports",
        exist_ok=True
    )

    file_path = os.path.join(
        "reports",
        f"progress_report_{learner_id}.xlsx"
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Progress Report"

    sheet.append(["Field", "Value"])
    sheet.append(["Learner ID", report["learner_id"]])
    sheet.append(["Lessons Completed", report["lessons_completed"]])
    sheet.append(["Average Accuracy", report["average_accuracy"]])
    sheet.append(["Practice Hours", report["practice_hours"]])
    sheet.append(["Improvement Rate", report["improvement_rate"]])
    sheet.append(["Current Streak", report["current_streak"]])
    sheet.append(["Certificates Earned", report["certificates_earned"]])

    workbook.save(file_path)

    return file_path


def export_class_summary_csv(
    db: Session,
) -> str:
    """
    Exports the instructor class summary as a CSV file.

    Returns:
        Path to the generated CSV file.
    """

    class_summary = (
        db.query(
            User,
            AnalyticsSummary,
        )
        .join(
            AnalyticsSummary,
            AnalyticsSummary.user_id == User.id,
        )
        .all()
    )

    os.makedirs(
        "reports",
        exist_ok=True
    )

    file_path = os.path.join(
        "reports",
        "class_summary_report.csv"
    )

    with open(
        file_path,
        mode="w",
        newline="",
        encoding="utf-8",
    ) as csv_file:

        writer = csv.writer(csv_file)

        writer.writerow(
            [
                "Learner ID",
                "Learner Name",
                "Lessons Completed",
                "Average Score",
                "Current Streak",
            ]
        )

        for user, analytics in class_summary:

            streak = streak_service.get_user_streak(
                db,
                str(user.id),
            )

            writer.writerow(
                [
                    str(user.id),
                    user.username,
                    analytics.lessons_completed,
                    analytics.overall_accuracy_percentage,
                    streak["current_streak"],
                ]
            )

    return file_path


def generate_accuracy_report(
    db: Session,
    learner_id: str,
) -> Dict:
    analytics = get_learner_analytics(
        db,
        learner_id,
    )

    return {
        "learner_id": learner_id,
        "overall_accuracy": analytics["average_accuracy"],
        "weekly_accuracy": analytics["weekly_accuracy"],
        "weak_letters": analytics["weak_letters"],
        "weekly_weak_letters": analytics["weekly_weak_letters"],
    }


def export_accuracy_report_pdf(
    db: Session,
    learner_id: str,
) -> str:
    report = generate_accuracy_report(
        db,
        learner_id,
    )

    os.makedirs(
        "reports",
        exist_ok=True
    )

    file_path = os.path.join(
        "reports",
        f"accuracy_report_{learner_id}.pdf"
    )

    document = SimpleDocTemplate(file_path)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Accuracy Report", styles["Title"]))
    story.append(Spacer(1, 20))

    rows = [
        ["Field", "Value"],
        ["Learner ID", report["learner_id"]],
        ["Overall Accuracy", report["overall_accuracy"]],
        [
            "Weak Letters",
            ", ".join(report["weak_letters"]) if report["weak_letters"] else "None",
        ],
    ]

    for week, accuracy in report["weekly_accuracy"].items():
        rows.append([f"Weekly Accuracy - {week}", accuracy])

    for week, letters in report["weekly_weak_letters"].items():
        rows.append(
            [
                f"Weekly Weak Letters - {week}",
                ", ".join(letters) if letters else "None",
            ]
        )

    table = Table(rows)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    story.append(table)
    document.build(story)

    return file_path


def export_accuracy_report_excel(
    db: Session,
    learner_id: str,
) -> str:
    report = generate_accuracy_report(
        db,
        learner_id,
    )

    os.makedirs(
        "reports",
        exist_ok=True
    )

    file_path = os.path.join(
        "reports",
        f"accuracy_report_{learner_id}.xlsx"
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Accuracy Report"

    sheet.append(["Field", "Value"])
    sheet.append(["Learner ID", report["learner_id"]])
    sheet.append(["Overall Accuracy", report["overall_accuracy"]])
    sheet.append(
        [
            "Weak Letters",
            ", ".join(report["weak_letters"]) if report["weak_letters"] else "None",
        ]
    )

    for week, accuracy in report["weekly_accuracy"].items():
        sheet.append([f"Weekly Accuracy - {week}", accuracy])

    for week, letters in report["weekly_weak_letters"].items():
        sheet.append(
            [
                f"Weekly Weak Letters - {week}",
                ", ".join(letters) if letters else "None",
            ]
        )

    workbook.save(file_path)

    return file_path


def generate_certification_report(
    db: Session,
    learner_id: str,
) -> Dict:
    certificates = (
        db.query(Certificate)
        .filter(
            Certificate.user_id == learner_id
        )
        .all()
    )

    return {
        "learner_id": learner_id,
        "certificates": [
            {
                "certificate_id": certificate.id,
                "issued_date": certificate.issued_date,
                "overall_score": certificate.overall_score,
                "pdf_url": certificate.pdf_url,
            }
            for certificate in certificates
        ],
    }


def export_certification_report_pdf(
    db: Session,
    learner_id: str,
) -> str:
    report = generate_certification_report(
        db,
        learner_id,
    )

    os.makedirs(
        "reports",
        exist_ok=True
    )

    file_path = os.path.join(
        "reports",
        f"certification_report_{learner_id}.pdf"
    )

    document = SimpleDocTemplate(file_path)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Certification Report", styles["Title"]))
    story.append(Spacer(1, 20))

    rows = [
        ["Certificate ID", "Issued Date", "Overall Score", "PDF URL"]
    ]

    for certificate in report["certificates"]:
        rows.append(
            [
                certificate["certificate_id"],
                certificate["issued_date"],
                certificate["overall_score"],
                certificate["pdf_url"] or "N/A",
            ]
        )

    table = Table(rows)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    story.append(table)
    document.build(story)

    return file_path


def export_certification_report_excel(
    db: Session,
    learner_id: str,
) -> str:
    report = generate_certification_report(
        db,
        learner_id,
    )

    os.makedirs(
        "reports",
        exist_ok=True
    )

    file_path = os.path.join(
        "reports",
        f"certification_report_{learner_id}.xlsx"
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Certification Report"

    sheet.append(["Certificate ID", "Issued Date", "Overall Score", "PDF URL"])

    for certificate in report["certificates"]:
        sheet.append(
            [
                certificate["certificate_id"],
                certificate["issued_date"],
                certificate["overall_score"],
                certificate["pdf_url"] or "N/A",
            ]
        )

    workbook.save(file_path)

    return file_path


def generate_learning_report(
    db: Session,
    learner_id: str,
) -> Dict:
    analytics = get_learner_analytics(
        db,
        learner_id,
    )

    return {
        "learner_id": learner_id,
        "lessons_completed": analytics["lessons_completed"],
        "weekly_improvement_rate": analytics["weekly_improvement_rate"],
    }


def export_learning_report_pdf(
    db: Session,
    learner_id: str,
) -> str:
    report = generate_learning_report(
        db,
        learner_id,
    )

    os.makedirs(
        "reports",
        exist_ok=True
    )

    file_path = os.path.join(
        "reports",
        f"learning_report_{learner_id}.pdf"
    )

    document = SimpleDocTemplate(file_path)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Learning Report", styles["Title"]))
    story.append(Spacer(1, 20))

    rows = [
        ["Field", "Value"],
        ["Learner ID", report["learner_id"]],
        ["Lessons Completed", report["lessons_completed"]],
    ]

    for week, rate in report["weekly_improvement_rate"].items():
        rows.append(
            [
                f"Weekly Improvement Rate - {week}",
                rate if rate is not None else "N/A",
            ]
        )

    table = Table(rows)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    story.append(table)
    document.build(story)

    return file_path


def export_learning_report_excel(
    db: Session,
    learner_id: str,
) -> str:
    report = generate_learning_report(
        db,
        learner_id,
    )

    os.makedirs(
        "reports",
        exist_ok=True
    )

    file_path = os.path.join(
        "reports",
        f"learning_report_{learner_id}.xlsx"
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Learning Report"

    sheet.append(["Field", "Value"])
    sheet.append(["Learner ID", report["learner_id"]])
    sheet.append(["Lessons Completed", report["lessons_completed"]])

    for week, rate in report["weekly_improvement_rate"].items():
        sheet.append(
            [
                f"Weekly Improvement Rate - {week}",
                rate if rate is not None else "N/A",
            ]
        )

    workbook.save(file_path)

    return file_path


def generate_assessment_report(
    db: Session,
    learner_id: str,
) -> Dict:
    assessments = (
        db.query(Assessment)
        .join(
            PracticeSession,
            Assessment.session_id == PracticeSession.id,
        )
        .filter(
            PracticeSession.user_id == learner_id
        )
        .all()
    )

    return {
        "learner_id": learner_id,
        "assessments": [
            {
                "assessment_id": assessment.id,
                "predicted_sign": assessment.predicted_sign,
                "expected_sign": assessment.expected_sign,
                "confidence": assessment.confidence,
                "hand_shape_score": assessment.hand_shape_score,
                "finger_position_score": assessment.finger_position_score,
                "timing_score": assessment.timing_score,
                "overall_accuracy": assessment.overall_accuracy,
                "is_correct": assessment.is_correct,
                "suggestions": assessment.suggestions,
                "created_at": assessment.created_at,
            }
            for assessment in assessments
        ],
    }


def export_assessment_report_pdf(
    db: Session,
    learner_id: str,
) -> str:
    report = generate_assessment_report(
        db,
        learner_id,
    )

    os.makedirs(
        "reports",
        exist_ok=True
    )

    file_path = os.path.join(
        "reports",
        f"assessment_report_{learner_id}.pdf"
    )

    document = SimpleDocTemplate(file_path)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Assessment Report", styles["Title"]))
    story.append(Spacer(1, 20))

    rows = [
        [
            "Predicted",
            "Expected",
            "Confidence",
            "Hand Shape",
            "Finger Position",
            "Timing",
            "Overall Accuracy",
            "Correct",
            "Created At",
        ]
    ]

    for assessment in report["assessments"]:
        rows.append(
            [
                assessment["predicted_sign"],
                assessment["expected_sign"],
                assessment["confidence"],
                assessment["hand_shape_score"],
                assessment["finger_position_score"],
                assessment["timing_score"],
                assessment["overall_accuracy"],
                assessment["is_correct"],
                assessment["created_at"],
            ]
        )

    table = Table(rows)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.darkblue),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )

    story.append(table)
    document.build(story)

    return file_path


def export_assessment_report_excel(
    db: Session,
    learner_id: str,
) -> str:
    report = generate_assessment_report(
        db,
        learner_id,
    )

    os.makedirs(
        "reports",
        exist_ok=True
    )

    file_path = os.path.join(
        "reports",
        f"assessment_report_{learner_id}.xlsx"
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Assessment Report"

    sheet.append(
        [
            "Predicted",
            "Expected",
            "Confidence",
            "Hand Shape",
            "Finger Position",
            "Timing",
            "Overall Accuracy",
            "Correct",
            "Created At",
        ]
    )

    for assessment in report["assessments"]:
        sheet.append(
            [
                assessment["predicted_sign"],
                assessment["expected_sign"],
                assessment["confidence"],
                assessment["hand_shape_score"],
                assessment["finger_position_score"],
                assessment["timing_score"],
                assessment["overall_accuracy"],
                assessment["is_correct"],
                str(assessment["created_at"]) if assessment["created_at"] else "",
            ]
        )

    workbook.save(file_path)

    return file_path